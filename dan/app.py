import os
import json
import csv
from flask import Flask, request, jsonify
from werkzeug.utils import secure_filename
from pathlib import Path
from dotenv import load_dotenv
from openpyxl import load_workbook
from openai import OpenAI

# Load environment variables
project_root = Path(__file__).resolve().parent
dan_folder = project_root / "dan"
env_path = dan_folder / ".env"
load_dotenv(dotenv_path=env_path)
llm_client = OpenAI(api_key=os.getenv("OPENAI_API_KEY"))

# Flask setup
app = Flask(__name__)
app.config['UPLOAD_FOLDER'] = 'uploads'
os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)

# Store processed data and extracted parameters
processed_files = {}

def process_csv(file_path):
    """Process a CSV file into JSON."""
    with open(file_path, 'r') as csvfile:
        reader = csv.DictReader(csvfile)
        return [row for row in reader]

def process_tsv(file_path):
    """Process a TSV file into JSON."""
    with open(file_path, 'r') as tsvfile:
        reader = csv.DictReader(tsvfile, delimiter='\t')
        return [row for row in reader]

def process_json(file_path):
    """Process a JSON file."""
    with open(file_path, 'r') as jsonfile:
        return json.load(jsonfile)

def process_xlsx(file_path):
    """Process an Excel (.xlsx) file into JSON."""
    workbook = load_workbook(file_path)
    sheet = workbook.active
    data = []
    headers = [cell.value for cell in sheet[1]]
    for row in sheet.iter_rows(min_row=2, values_only=True):
        data.append(dict(zip(headers, row)))
    return data

def determine_ranked_parameters(data):
    """Use LLM to determine the ranked parameters for a dataset."""
    sample_data = data[:5]  # Use the first 5 rows for analysis
    prompt = (
        f"Analyze the following dataset and rank the most important attributes from most to least important and only listing the name of the actual attribute and nothing else for each entry, "
        f"based on the titles of the metadata:\n\n{json.dumps(sample_data)}"
    )
    try:
        completion = llm_client.chat.completions.create(
            model="gpt-4o-mini",
            messages=[
                {"role": "system", "content": "You are a helpful assistant."},
                {"role": "user", "content": prompt}
            ]
        )
        result = completion.choices[0].message["content"].strip()
        return result.split(", ")  # Return a ranked list of attributes
    except Exception as e:
        print(f"Error ranking parameters: {e}")
        return []

@app.route('/upload', methods=['POST'])
def upload_files():
    """Handle file uploads and process them."""
    global processed_files
    if 'folder' not in request.files:
        return jsonify({"error": "No folder uploaded."}), 400

    files = request.files.getlist('folder')
    response = []

    for file in files:
        filename = secure_filename(file.filename)
        file_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
        os.makedirs(os.path.dirname(file_path), exist_ok=True)
        file.save(file_path)

        # Determine file type and process
        extension = os.path.splitext(filename)[1].lower()
        if extension in ['.csv', '.tsv', '.json', '.xlsx']:
            if extension == '.csv':
                processed_data = process_csv(file_path)
            elif extension == '.tsv':
                processed_data = process_tsv(file_path)
            elif extension == '.json':
                processed_data = process_json(file_path)
            elif extension == '.xlsx':
                processed_data = process_xlsx(file_path)

            # Store processed data and determine ranked parameters
            processed_files[filename] = processed_data
            ranked_parameters = determine_ranked_parameters(processed_data)
            response.append({
                "file_name": filename,
                "ranked_parameters": ranked_parameters
            })
        else:
            response.append({
                "file_name": filename,
                "error": "Unsupported file format"
            })

    return jsonify({"files": response})

@app.route('/parameters', methods=['GET'])
def get_parameters():
    """Return the ranked parameters for each file."""
    if not processed_files:
        return jsonify({"error": "No files have been processed yet."}), 400

    response = []
    for file_name, data in processed_files.items():
        ranked_parameters = determine_ranked_parameters(data)
        response.append({
            "file_name": file_name,
            "ranked_parameters": ranked_parameters
        })

    return jsonify({"files": response})

if __name__ == '__main__':
    app.run(debug=True)
