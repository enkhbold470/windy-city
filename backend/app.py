import os
import json
import csv
from flask import Flask, request, jsonify
from werkzeug.utils import secure_filename
from pathlib import Path
from dotenv import load_dotenv
from openpyxl import load_workbook
from flask_cors import CORS


# Load environment variables
project_root = Path(__file__).resolve().parent
dan_folder = project_root / "dan"
env_path = dan_folder / ".env"
load_dotenv(dotenv_path=env_path)

# Flask setup
app = Flask(__name__)
CORS(app)  # Enable CORS for all routes
app.config['UPLOAD_FOLDER'] = 'uploads'
os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)

# Helper functions
def process_csv(file_path):
    """Process a CSV file into JSON."""
    with open(file_path, 'r') as csvfile:
        reader = csv.DictReader(csvfile)
        return [row for row in reader]

def process_tsv(file_path):
    """Process a TSV file into JSON."""
    with open(file_path, 'r') as tsvfile:
        reader = csv.DictReader(tsvfile, delimiter='\t')
        return [row for row in reader]

def process_json(file_path):
    """Process a JSON file."""
    with open(file_path, 'r') as jsonfile:
        return json.load(jsonfile)

def process_xlsx(file_path):
    """Process an Excel (.xlsx) file into JSON."""
    workbook = load_workbook(file_path)
    sheet = workbook.active
    data = []
    headers = [cell.value for cell in sheet[1]]
    for row in sheet.iter_rows(min_row=2, values_only=True):
        data.append(dict(zip(headers, row)))
    return data

def determine_ranked_parameters(data):
    """
    Rank parameters based on the total sum of numerical values for each attribute.
    Parameters with the highest total sum of values come first.
    """
    # Initialize a dictionary to store the sum of numerical values for each attribute
    attribute_sums = {}

    # Iterate over the data rows
    for row in data:
        for key, value in row.items():
            # Check if the value is numeric (float or int)
            try:
                num_value = float(value)  # Try converting to a float
                attribute_sums[key] = attribute_sums.get(key, 0) + num_value
            except (ValueError, TypeError):
                continue  # Ignore non-numeric values

    # Sort attributes by the total sum of their numerical values in descending order
    sorted_attributes = sorted(attribute_sums.items(), key=lambda item: item[1], reverse=True)
    
    # Extract just the attribute names in sorted order
    ranked_attributes = [attribute for attribute, _ in sorted_attributes]
    return ranked_attributes

@app.route('/process', methods=['POST'])
def process_files():
    """Handle file uploads, process them, and return results."""
    if 'files' not in request.files:
        return jsonify({"error": "No files uploaded."}), 400

    files = request.files.getlist('files')
    results = []

    for file in files:
        filename = secure_filename(file.filename)
        file_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
        file.save(file_path)

        extension = os.path.splitext(filename)[1].lower()
        if extension == '.csv':
            processed_data = process_csv(file_path)
        elif extension == '.xlsx':
            processed_data = process_xlsx(file_path)
        else:
            results.append({"file_name": filename, "error": "Unsupported file format"})
            continue

        ranked_parameters = determine_ranked_parameters(processed_data)

        # Ensure the processed data is JSON serializable
        cleaned_data = [
            {key: value for key, value in row.items() if key is not None and value is not None}
            for row in processed_data
        ]

        # Append results
        results.append({
            "file_name": filename,
            "ranked_parameters": ranked_parameters,
            "processed_data": cleaned_data
        })

    # Clean response data to ensure JSON serializability
    clean_results = []
    for result in results:
        clean_result = {key: value for key, value in result.items() if key is not None}
        if "processed_data" in clean_result:
            clean_result["processed_data"] = [
                {k: v for k, v in row.items() if k is not None and v is not None}
                for row in clean_result["processed_data"]
            ]
        clean_results.append(clean_result)

    return jsonify({
        "results": clean_results,
        "message": "Ha Inky I got it to work (:"
    })

if __name__ == '__main__':
    app.run(debug=True)
